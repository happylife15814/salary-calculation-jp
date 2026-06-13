"""
給与計算プログラム - 2026年5月分
"""

import os
import unicodedata
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# ===== 設定 =====
YEAR_MONTH = '2026年05月'
UPLOAD_DIR = '/home/ubuntu/upload'
OUTPUT_DIR = '/home/ubuntu/salary_output_may'
TEMPLATE_FILE = os.path.join(UPLOAD_DIR, '給与明細フォーマット_blank.xlsx')
KINTAI_FILE = os.path.join(UPLOAD_DIR, '【2026年05月】勤務表.xlsx')
DAILY_MASTER = os.path.join(UPLOAD_DIR, 'シフト日常清掃.xlsx')
KAJI_MASTER = os.path.join(UPLOAD_DIR, 'シフト家事時間.xlsx')
TRAINING_MASTER = os.path.join(UPLOAD_DIR, 'シフト研修と引継ぎ.xlsx')
MTG_MASTER = os.path.join(UPLOAD_DIR, 'シフトMTG.xlsx')
TRANSPORT_MASTER = os.path.join(UPLOAD_DIR, '交通費表（担当者別）.xlsx')
HEIKIN_MASTER = os.path.join(UPLOAD_DIR, '過去3ヵ月の平均勤務時間.xlsx')
KAGI_MASTER = os.path.join(UPLOAD_DIR, '鍵預かりお礼金リスト.xlsx')

os.makedirs(OUTPUT_DIR, exist_ok=True)


def normalize(s):
    """全角→半角変換 + 空白除去"""
    if s is None:
        return ''
    s = str(s).strip()
    s = unicodedata.normalize('NFKC', s)
    s = s.replace(' ', '').replace('\u3000', '')
    return s


def load_kagi_list():
    """鍵預かりお礼金対象者リストをファイルから読み込む"""
    kagi = set()
    wb = load_workbook(KAGI_MASTER)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0]:
            kagi.add(normalize(row[0]))
    wb.close()
    print(f'鍵預かりお礼金対象者: {sorted(kagi)}')
    return kagi


def load_wages():
    """全マスターのシフト→時給辞書を読み込む"""
    wages = {
        'daily': {},
        'kaji': set(),
        'training': {},
        'mtg': set(),
    }

    # 日常清掃
    wb = load_workbook(DAILY_MASTER)
    ws = wb['日常清掃時給表']
    for row in ws.iter_rows(min_row=2, max_row=50, min_col=1, max_col=2, values_only=True):
        name, wage = row
        if name and name != 'シフト名' and wage:
            wages['daily'][normalize(name)] = wage
    wb.close()

    # 家事時間
    wb = load_workbook(KAJI_MASTER)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, values_only=True):
        for cell in row:
            if cell:
                wages['kaji'].add(normalize(cell))
    wb.close()

    # 研修
    wb = load_workbook(TRAINING_MASTER)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, min_col=1, max_col=2, values_only=True):
        name, wage = row[0], row[1]
        if name and name not in ('研修と引継ぎ', 'シフト名'):
            wages['training'][normalize(name)] = wage if wage else 1150
    wb.close()

    # MTG
    wb = load_workbook(MTG_MASTER)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0]:
            wages['mtg'].add(normalize(row[0]))
    wb.close()

    return wages


def get_shift_category(shift_name, wages):
    """シフト名からカテゴリを返す"""
    n = normalize(shift_name)
    if n in wages['daily']:
        return 'daily', wages['daily'][n]
    if n in wages['training']:
        return 'training', wages['training'][n]
    if n in wages['mtg']:
        return 'mtg', 1800
    if n in wages['kaji']:
        return 'kaji', None
    return 'unknown', None


def load_heikin():
    """過去3ヵ月平均勤務時間を読み込む"""
    heikin = {}
    wb = load_workbook(HEIKIN_MASTER)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        m1 = row[1]
        m2 = row[2]
        m3 = row[3]
        avg = row[4]
        if name:
            if isinstance(avg, (int, float)):
                avg_val = float(avg)
            else:
                vals = [v for v in [m1, m2, m3] if isinstance(v, (int, float))]
                avg_val = sum(vals) / 3 if vals else 0.0
            heikin[normalize(name)] = {
                'months': [m1, m2, m3],
                'avg': avg_val
            }
    wb.close()
    return heikin


def get_kaji_wage(avg_hours):
    """過去3ヵ月平均時間から家事時間の時給を決定"""
    if avg_hours >= 40:
        return 1800
    elif avg_hours >= 12:
        return 1700
    else:
        return 1600


def load_transport(name):
    """交通費マスターから対象者の交通費を読み込む"""
    transport = {}
    wb = load_workbook(TRANSPORT_MASTER)
    ws = wb.active
    target_found = False
    name_norm = normalize(name).replace('様', '')

    for row in ws.iter_rows(min_row=2, max_row=300, min_col=1, max_col=4, values_only=True):
        name_cell, shift_cell, _, fee_cell = row[0], row[1], row[2], row[3]

        if name_cell:
            master_name_norm = normalize(name_cell).replace('様', '')
            if name_norm in master_name_norm or master_name_norm in name_norm:
                target_found = True
            else:
                if target_found:
                    break
                target_found = False

        if target_found and shift_cell and fee_cell is not None:
            transport[normalize(shift_cell)] = fee_cell

    wb.close()
    return transport


def process_kintai(ws_kintai, wages, kaji_wage):
    """勤務表シートから勤務記録を解析する"""
    records = []
    skip_shifts = {'法定休日', '所定休日', '有給休暇', '欠勤', '休日', ''}

    for row in ws_kintai.iter_rows(min_row=11, max_row=50, values_only=True):
        date_val = row[0] if len(row) > 0 else None
        shift_val = row[1] if len(row) > 1 else None
        hours_val = row[10] if len(row) > 10 else None

        if not date_val or not shift_val:
            continue
        shift_str = str(shift_val).strip()
        if normalize(shift_str) in {normalize(s) for s in skip_shifts}:
            continue

        if isinstance(hours_val, (int, float)):
            hours = float(hours_val)
        elif isinstance(hours_val, str):
            try:
                hours = float(hours_val)
            except ValueError:
                hours = 0.0
        else:
            hours = 0.0

        category, wage = get_shift_category(shift_str, wages)

        if category == 'kaji':
            wage = kaji_wage
        elif category == 'mtg':
            wage = 1800
            hours = 0.0
        elif category == 'unknown':
            wage = 0

        payment = hours * wage if category not in ('mtg', 'unknown') else (1800 if category == 'mtg' else 0)

        records.append({
            'date': date_val,
            'shift': shift_str,
            'category': category,
            'hours': hours,
            'wage': wage or 0,
            'payment': payment,
            'row_data': row,
        })

    return records


def write_output(name, records, heikin_info, transport_fees, kagi_list):
    """給与明細を作成して保存する"""
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    # ヘッダー
    ws['A3'] = name
    months = heikin_info.get('months', [None, None, None])
    ws['I3'] = months[0] if months[0] is not None else 0
    ws['J3'] = months[1] if months[1] is not None else 0
    ws['K3'] = months[2] if months[2] is not None else 0
    ws['E3'] = '=(I3+J3+K3)/3'

    # 明細行（12行目～）
    row_num = 12
    for rec in records:
        if rec['category'] == 'mtg':
            continue
        r = rec['row_data']

        # A列: 日付, B列: シフト名
        ws[f'A{row_num}'] = rec['date']
        ws[f'B{row_num}'] = rec['shift']

        # C列～K列を転記
        for col_idx, col_letter in enumerate(['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'], start=2):
            val = r[col_idx] if col_idx < len(r) else None
            ws[f'{col_letter}{row_num}'] = val

        # L列: 支給額
        if rec['payment'] > 0:
            ws[f'L{row_num}'] = rec['payment']
            ws[f'L{row_num}'].number_format = '#,##0'

        # M列: 交通費
        shift_norm = normalize(rec['shift'])
        transport_fee = transport_fees.get(shift_norm, 0)
        if transport_fee:
            ws[f'M{row_num}'] = transport_fee
            ws[f'M{row_num}'].number_format = '#,##0'

        row_num += 1

    # 集計欄
    kaji_records = [r for r in records if r['category'] == 'kaji']
    daily_records = [r for r in records if r['category'] == 'daily']
    training_records = [r for r in records if r['category'] == 'training']
    mtg_records = [r for r in records if r['category'] == 'mtg']

    kaji_hours = sum(r['hours'] for r in kaji_records)
    kaji_payment = sum(r['payment'] for r in kaji_records)
    daily_hours = sum(r['hours'] for r in daily_records)
    daily_payment = sum(r['payment'] for r in daily_records)
    training_hours = sum(r['hours'] for r in training_records)
    training_payment = sum(r['payment'] for r in training_records)
    mtg_count = len(mtg_records)
    mtg_payment = mtg_count * 1800

    # 交通費合計
    total_transport = 0
    for rec in records:
        if rec['category'] == 'mtg':
            continue
        shift_norm = normalize(rec['shift'])
        total_transport += transport_fees.get(shift_norm, 0)

    # 鍵預かりお礼金
    kagi_fee = 500 if normalize(name) in kagi_list else 0

    # 総支給額
    total_payment = kaji_payment + daily_payment + training_payment + mtg_payment + kagi_fee

    # J5: 家事時間 合計時間, L5: 家事時間 合計支給額
    if kaji_hours > 0:
        ws['J5'] = kaji_hours
        ws['J5'].number_format = '0.0'
        ws['L5'] = kaji_payment
        ws['L5'].number_format = '#,##0'

    # M5: 日常清掃 合計時間, O5: 日常清掃 合計支給額
    if daily_hours > 0:
        ws['M5'] = daily_hours
        ws['M5'].number_format = '0.0'
        ws['O5'] = daily_payment
        ws['O5'].number_format = '#,##0'

    # M7: 研修 合計時間, N7: 研修時給, O7: 研修合計支給額
    if training_hours > 0:
        ws['M7'] = training_hours
        ws['M7'].number_format = '0.0'
        ws['N7'] = 1150
        ws['N7'].number_format = '#,##0'
        ws['O7'] = training_payment
        ws['O7'].number_format = '#,##0'

    # D5: MTG手当
    if mtg_payment > 0:
        ws['D5'] = mtg_payment
        ws['D5'].number_format = '#,##0'
        print(f'  MTG手当: {mtg_payment:,}円 → D5セル')

    # I5: 鍵預かりお礼金
    if kagi_fee > 0:
        ws['I5'] = kagi_fee
        ws['I5'].number_format = '#,##0'
        print(f'  鍵預かりお礼金: {kagi_fee}円 → I5セル')

    # A5: 総支給額（計算式を上書きして数値として書き込む）
    # = 家事時間(L5) + 交通費(M33) + 日常清掃(O5) + 研修(O7) + MTG(D5) + 鍵預かり(I5) + その他手当
    ws['A5'] = total_payment + total_transport
    ws['A5'].number_format = '#,##0'

    # C5: 交通費合計（計算式を上書きして数値として書き込む）
    ws['C5'] = total_transport
    ws['C5'].number_format = '#,##0'

    # 書式設定
    center_align = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_align

    for row_num_f in range(1, 10):
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row_num_f, col_num)
            cell.font = Font(size=12, bold=False)

    for row_num_f in range(11, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row_num_f, col_num)
            cell.font = Font(size=10, bold=False)

    for r in range(12, ws.max_row + 1):
        cell = ws[f'K{r}']
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '0.00'

    # 保存
    out_path = os.path.join(OUTPUT_DIR, f'給与明細_{YEAR_MONTH}_{name}.xlsx')
    wb.save(out_path)
    wb.close()

    return total_payment, total_transport


def main():
    print('マスターファイルを読み込み中...')
    wages = load_wages()
    heikin = load_heikin()
    kagi_list = load_kagi_list()

    wb_kintai = load_workbook(KINTAI_FILE)
    count = 0
    unknown_shifts = {}

    print(f'\n処理開始: {YEAR_MONTH}')
    print('=' * 60)

    for sheet_name in wb_kintai.sheetnames:
        ws = wb_kintai[sheet_name]
        name = ws['E3'].value
        if not name:
            continue

        name = str(name).strip()
        print(f'\n処理中: {name}')

        name_norm = normalize(name)
        heikin_info = heikin.get(name_norm, {'months': [0, 0, 0], 'avg': 0.0})
        avg_hours = heikin_info['avg']
        kaji_wage = get_kaji_wage(avg_hours)
        print(f'  過去3ヵ月平均: {avg_hours:.2f}時間 → 家事時給: {kaji_wage}円')

        transport_fees = load_transport(name)
        records = process_kintai(ws, wages, kaji_wage)
        print(f'  勤務記録: {len(records)}件')

        for rec in records:
            if rec['category'] == 'unknown':
                shift = rec['shift']
                if shift not in unknown_shifts:
                    unknown_shifts[shift] = []
                unknown_shifts[shift].append(name)
                print(f'  警告: シフト「{shift}」がマスターに見つかりません')

        total_payment, total_transport = write_output(name, records, heikin_info, transport_fees, kagi_list)
        print(f'  ✓ 完了: 給与明細_{YEAR_MONTH}_{name}.xlsx（総支給額: {total_payment:,.0f}円 + 交通費: {total_transport:,.0f}円）')
        count += 1

    wb_kintai.close()

    print('\n' + '=' * 60)
    print(f'作成した給与明細: {count}件')
    if unknown_shifts:
        print('\n【要確認】マスターに未登録のシフト:')
        for shift, names in unknown_shifts.items():
            print(f'  「{shift}」 - 対象者: {", ".join(set(names))}')
    else:
        print('未登録シフト: なし')
    print('=' * 60)


if __name__ == '__main__':
    main()
